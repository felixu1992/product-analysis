import os
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook


class Product:

    def __init__(self):
        """
        构造 ProductInfo 对象，后续所有 Excel 数据均解析为此对象
        """

class ProductGroup:
    def __init__(self, name, products):
        self.name = name
        self.products = products
        """
        构造 ProductGroup 对象，用以对 Product 进行按 product_name 进行分组
        """


class ExcelParser:
    def __init__(self, filename):
        """
        持有整个 Excel
        """
        self.work_book = load_workbook(filename)

    def get_sheet_names(self):
        """
        拿到当前 Excel 所有的 Sheet
        """
        return self.work_book.sheetnames

    def get_sheet_count(self):
        """
        拿到当前 Excel Sheet 的总数
        """
        return len(self.work_book.sheetnames)

class ExcelWriter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "年度报告"

    def create_excel_report(self, filename):
        # 定义样式
        header_font = Font(bold=True, color="0070C0")
        highlight_fill = PatternFill("solid", fgColor="FFFF00")

        # 构建表头
        headers = ["部门", "Q1", "Q2", "Q3", "Q4", "年平均"]
        self.ws.append(headers)

        # 设置标题样式
        for col in range(1, 7):
            cell = self.ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        departments = ["市场部", "技术部", "财务部"]
        data = [
            [850, 920, 880, 950],
            [1200, 1350, 1420, 1500],
            [680, 720, 700, 750]
        ]

        for i, dept in enumerate(departments):
            row = [dept] + data[i] + [f"=AVERAGE(B{i + 2}:E{i + 2})"]
            self.ws.append(row)

        # 设置数字格式
        for row in self.ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=6):
            for cell in row:
                cell.number_format = "#,##0"

        # 获取当前目录的父目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        parent_dir = os.path.dirname(current_dir)

        # 构建目标路径
        target_dir = os.path.join(parent_dir, "report")
        os.makedirs(target_dir, exist_ok=True)

        # 保存文件
        file_path = os.path.join(target_dir, filename)
        self.wb.save(file_path)


if __name__ == '__main__':
    writer = ExcelWriter()
    writer.create_excel_report("abc.xlsx")


class ProductHolder:

    def __init__(self, work_book, sheet_name):
        """
        解析当前 sheet 页为 Product 对象列表

        :param work_book: 整个工作簿
        :param sheet_name: 当前 sheet 页
        """
        # 取工作簿中对应 sheet 页保存
        self.sheet = work_book[sheet_name]
        # 获取数据总行数
        self.rows = list(self.sheet.rows)
        # 获取表头
        self.title = [column.value for column in self.rows[0]]
        # 产品列表
        self.products = []
        # 构建所有产品信息
        self.read_product()
        # 产品组
        self.product_groups = []
        # 构建产品组
        self.group_by_product_code()

    def read_product(self):
        """
        构建所有产品信息
        """
        project_count = len(self.rows)
        # 从第 2 行开始取产品信息
        for row in range(1, project_count):
            # 获取整行信息为列表
            info = [column.value for column in self.rows[row]]
            # 定义当前要处理的用例对象
            product = Product()
            # 将表头字段对应的列作为对象字段，将当前行对应的列信息作为值，填入产品对象中
            for i in zip(self.title, info):
                setattr(product, i[0], i[1])
            # 定义当前用例所在行数(其实该行数为实际对应 Excel 中行数减 1)
            product.row = row
            # case_info.method = case_info.method.lower()
            # 加入用例列表
            self.products.append(product)

    def group_by_product_code(self):
        """
        对所有的产品信息进行分组
        """
        # 使用 defaultdict 进行分组
        group_dict = defaultdict(list)
        # 根据 product_line 分组
        for product in self.products:
            group_dict[product.product_name].append(product)
        # 转换为 ProductGroup 对象列表
        self.product_groups = [
            ProductGroup(name=name, products=products)
            for name, products in group_dict.items()
        ]